import telebot  # библиотека для разработки telegram-ботов
from telebot import types  # библиотека для создания кнопок
import pandas as pd  # высокоуровневая библиотека для анализа данных
from openpyxl import load_workbook  # библиотека для работы с файлами Excel
import requests  # библиотека запросов
from bs4 import BeautifulSoup as bs  # библиотека запросов для извлечения данных из файлов HTML и XML
from copy import copy  # метод поверхностное и глубокое копирование объектов
from time import sleep  # для симуляции задержки в выполнении программы


def copy_cell(src_sheet, src_row, src_col,  # копирование строчки из одного Exel в конец другого
              tgt_sheet, tgt_row, tgt_col,
              copy_style=True):
    cell = src_sheet.cell(src_row, src_col)
    new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
    if cell.has_style and copy_style:
        new_cell._style = copy(cell._style)


source_currency = 'EUR'
destination_currency = 'BYN'


def convert_currency_xe(src, dst, cost):  # функция конвертер валют
    def get_digits(text):
        new_text = ""
        for c in text:
            if c.isdigit() or c == ".":
                new_text += c
        return float(new_text)

    url = f"https://www.xe.com/currencyconverter/convert/?Amount={cost}&From={src}&To={dst}"
    content = requests.get(url).content
    soup = bs(content, "html.parser")  #
    exchange_rate_html = soup.find_all("p")[2]

    return get_digits(exchange_rate_html.text)


bot = telebot.TeleBot('5723544785:AAGcG2-Fui41fp1Sl0E1hzrZGh4VEzZnUdE')  # создали бота

file = 'склад.xlsx'  # файл с данными по складу


df = pd.read_excel(file, index_col='ID')  # читаем Excel файл в DataFrame
pd.set_option('display.max_rows', None)  # без ограничений
pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)

df['№'], df['Name'] = df['№ and Name'].str.split(' ', 1).str  # разделяем столбец с каталожным номером и описанием
df = df.drop(columns='№ and Name')
df_new = df.reindex(columns=['№', 'Name', 'Amount', 'Price'])  # переименовываем


@bot.message_handler(content_types=['text'])  # объявили метод для получения текстовых сообщений
def get_user_text(message):  # старт бота
    if message.text == '/start':
        msg = bot.send_message(message.chat.id, f'Здравствуйте, {message.from_user.first_name}. Введите и отправьте'
                                                f' каталожный номер запрашиваемой детали. Номер может состоять из'
                                                f' цифр и латинских букв, пример B2424210000. Или Вы его можете найти в'
                                                f' Parts List на ваш класс оборудования, пример можете увидеть ниже ;-)')
        img = open('Пример каталожного номера.PNG', 'rb')
        bot.send_photo(message.chat.id, img)  # пример номера детали
        bot.register_next_step_handler(msg, create)  # запоминаем ввод для след функции


def create(message):  # создаем информацию по детали
    try:
        bot.send_message(message.chat.id, 'Идет проверка № детали\n{}'.format(message.text))
        need = df_new.loc[df_new['№'] == message.text]  # выбираем строку, соответствующую номеру детали
        writer = pd.ExcelWriter('output.xlsx')  # записываем строку в output.xlsx
        need.to_excel(writer)
        writer.save()  # сохраняем запись
        rezerv = pd.ExcelWriter('rezerv.xlsx')  # записываем сразу в бронь и будем там менять кол-во
        need.to_excel(rezerv)
        rezerv.save()
        workbook = load_workbook('rezerv.xlsx') # принимает имя файла в качестве аргумента и возвращает объект рабочей книги, который представляет файл.
        sheet = workbook.active  # служит для чтения значений свойств
        sheet["F1"] = "email"
        workbook.save('rezerv.xlsx')
        isempty = need.empty  # проверяем нашло ли деталь(если пустая то True)
        if not isempty:  # если совпадение есть, то
            df_cr = pd.read_excel('output.xlsx', index_col='ID')
            bot.send_message(message.chat.id, f"№ детали введен верно, это {df_cr.iloc[0]['Name']}")
            markup = types.InlineKeyboardMarkup(row_width=1)  # создаем inline кнопки
            but1 = types.InlineKeyboardButton("Количество на складе", callback_data='Amount')
            but2 = types.InlineKeyboardButton("Стоимость", callback_data='Price')
            markup.add(but1, but2)  # добавляем кнопки
            bot.send_message(message.chat.id, 'Что хотите узнать?', reply_markup=markup)
            sleep(8)
            msg = bot.send_message(message.chat.id, 'Сколько деталей хотите заказать?')
            bot.register_next_step_handler(msg, reservation)
        else:
            bot.send_message(message.chat.id, 'Такой детали нет в наличии или № детали введен неверно')
    except Exception:
        bot.reply_to(message, 'oooops')


def reservation(message):  # записываем количество деталей для заказа и итоговую сумму
    try:
        df_res = pd.read_excel('output.xlsx', index_col='ID')
        if message.text.isdigit():
            workbook = load_workbook('rezerv.xlsx')  # для открытия excel-файла для чтения
            sheet = workbook.active
            summa = float(df_res.iloc[0]['Price'])
            amount = float(message.text)
            exchange_rate = convert_currency_xe(source_currency, destination_currency, summa)
            exchange_rate = round(exchange_rate, 2)  # посчитали стоимость по курсу одной детали
            sheet["D2"] = message.text  # записали кол-во для заказа
            sheet["E2"] = amount * exchange_rate  # посчитали итоговую сумму
            workbook.save('rezerv.xlsx')
        else:
            bot.send_message(message.chat.id, 'Данные введены неверно')
        msg = bot.send_message(message.chat.id, 'Введите почту для связи по заказу')
        bot.register_next_step_handler(msg, info)
    except Exception:
        bot.reply_to(message, 'oooops')


def info(message):  # запись данных клиента и вывод заказа
    try:
        workbook = load_workbook('rezerv.xlsx')
        sheet = workbook.active
        sheet["F2"] = message.text  # записываем почту
        workbook.save('rezerv.xlsx')
        df_inf = pd.read_excel('rezerv.xlsx')
        bot.send_message(message.chat.id, f"Заказ деталей № {df_inf.iloc[0]['№']} - {df_inf.iloc[0]['Name']} \n"
                                          f"Количество {df_inf.iloc[0]['Amount']}\n"
                                          f"Сумма - {round(df_inf.iloc[0]['Price'], 2)} BYN",
                                          parse_mode='html')
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        but1 = types.KeyboardButton("Заказать")
        but2 = types.KeyboardButton("Отмена")
        markup.add(but1, but2)  # создаем Callback кнопки
        msg = bot.send_message(message.chat.id, "Нажмите заказать, если данные верны", reply_markup=markup)
        bot.register_next_step_handler(msg, order)
    except Exception:
        bot.reply_to(message, 'oooops')


def order(message):  # подтверждение заказа
    try:
        if message.text == "Заказать":
            df_reed = pd.read_excel('rezerv.xlsx')
            wb = load_workbook('rezerv.xlsx')
            wb1 = load_workbook('Заказы.xlsx')
            ws2 = wb.active
            ws1 = wb1.active
            ws1_last_row = ws1.max_row  # находим последнюю заполненную строку
            for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=2), 1):  # переносим вторую строку из rezerv.xlsx в конец в Заказы
                for cell in row:
                    copy_cell(ws2, cell.row, cell.column,
                              ws1, ws1_last_row + i, cell.column)
            wb1.save('Заказы.xlsx')
            bot.send_message(message.chat.id, f"Ваш заказ передан в обработку, ждите сообщения на {df_reed.iloc[0]['email']}\n"
                                              f"Для следующего заказа нажмите → /start")
        elif message.text == "Отмена":
            bot.send_message(message.chat.id, "Вы отменили заказ, нажмите → /start для повторного использования")
    except Exception:
        bot.reply_to(message, 'oooops')


@bot.callback_query_handler(func=lambda call: True)  # функционал кнопок информации
def callback(call):
    df_clb = pd.read_excel('output.xlsx', index_col='ID')
    cost = float(df_clb.iloc[0]['Price'])
    exchange_rate = convert_currency_xe(source_currency, destination_currency, cost)  # перевод в бел рубли с округлением
    exchange_rate = round(exchange_rate, 2)
    if call.message:
        if call.data == 'Amount':
            bot.send_message(call.from_user.id, f"Количество деталей № {df_clb.iloc[0]['№']} - {df_clb.iloc[0]['Amount']}",
                             parse_mode='html')
        if call.data == 'Price':
            bot.send_message(call.from_user.id, f"Стоимость детали № {df_clb.iloc[0]['№']} - {exchange_rate} BYN",
                             parse_mode='html')


bot.polling(none_stop=True)
